"""EIA 원본 두 개를 대상 월의 일별 연료비로 정제합니다.

산출물은 Gold 가 읽는 `gas_ev_price` Silver 입니다.
Gold 는 어느 경로로 만들어졌는지 몰라도 되고, 구분이 필요하면 `price_source` 를 봅니다.

정제에서 하는 일
---------------
1. 대상 월 추출 — 이력 파일에서 그 달에 해당하는 구간만
2. 일별 전개 — 휘발유는 주간, 전력은 월간이라 그 값이 유효한 날에 복제
3. 단위 변환 — 전력은 ¢/kWh 로 오므로 $/kWh 로
4. 마진 보정 — 전력 소매요금에 공공 충전 배수를 곱해 충전 단가로

3·4 를 거치는 이유
-----------------
휘발유는 AAA 와 EIA 가 **같은 것**(주유소 소매가)을 재므로 실측을 그대로 씁니다.
전기는 다릅니다 — EIA 는 전력 소매요금, NLR 은 공공 충전소 요금이고 실측상 2배쯤
차이 납니다. 그대로 넣으면 전기차 연료비가 절반으로 잡히므로 배수를 곱합니다.

pandas 를 쓰지 않는 이유
----------------------
이 런타임에는 pandas 가 없습니다(`pyproject.toml` — Lambda 이미지 크기와 glibc 상한
때문). 셀을 읽는 것이 전부라 `xlrd`/`openpyxl` 로 충분합니다.
"""

import calendar
import logging
from datetime import date, datetime
from io import BytesIO

import openpyxl
import xlrd

from schema.silver.gas_ev_price import EIA as PRICE_SOURCE, FINAL

logger = logging.getLogger(__name__)

# 휘발유 이력 시트 — 0번은 목차, 1번이 주간 계열입니다.
GAS_SHEET_INDEX = 1
GAS_HEADER_ROWS = 3
# 전력 이력 시트 — 부문별 요금이 2단 헤더(부문 / 항목)로 들어 있습니다.
ELECTRICITY_SHEET = "Monthly-States"
ELECTRICITY_HEADER_ROWS = 3
STATE = "NY"
# 전기차 충전은 교통 부문으로 분류됩니다. 가정용(RESIDENTIAL)은 자가 충전 기준이라
# 영업용 기사의 단가로 쓰기에 맞지 않습니다.
ELECTRICITY_SECTOR = "TRANSPORTATION"
# EIA 가 각 달의 확정 여부를 직접 적어주는 컬럼 (Preliminary / Final).
STATUS_COLUMN = "Data Status"

# 공공 충전 마진 배수. NLR 실측 공공 충전 단가 / EIA 전력 소매요금 으로 잰 값입니다
# (2026-08 기준 $0.417 / $0.207). **가정입니다** — 마진율이 시간에 따라 크게 변하지
# 않는다고 보고 과거에 같은 배수를 적용합니다. 재려면 두 값을 같은 시점에 다시 보세요.
PUBLIC_CHARGING_MARKUP = 2.0
CENTS_PER_DOLLAR = 100.0

# 이상값 차단. 뉴욕 소매가가 이 범위를 벗어나면 원본 형식이 바뀐 것으로 봅니다.
GAS_USD_RANGE = (1.0, 15.0)
EV_USD_RANGE = (0.05, 3.0)

# `price_source` / 확정 상태 값은 스키마가 소유합니다 (위 import). 두 곳에 적어두면
# 갈릴 수 있어서 여기서 다시 정의하지 않습니다.


def month_days(year_month: str) -> list[date]:
    year, month = (int(part) for part in year_month.split("-"))
    last = calendar.monthrange(year, month)[1]
    return [date(year, month, day) for day in range(1, last + 1)]


def parse_gas_weekly(body: bytes) -> list[tuple[date, float]]:
    """주간 휘발유 이력 → (관측일, USD/gal) 오름차순."""
    book = xlrd.open_workbook(file_contents=body)
    if book.nsheets <= GAS_SHEET_INDEX:
        raise ValueError("EIA 휘발유 파일에 주간 계열 시트가 없습니다")
    sheet = book.sheet_by_index(GAS_SHEET_INDEX)

    observations: list[tuple[date, float]] = []
    for index in range(GAS_HEADER_ROWS, sheet.nrows):
        raw_date, raw_price = sheet.cell_value(index, 0), sheet.cell_value(index, 1)
        if not isinstance(raw_date, float) or not isinstance(raw_price, float):
            continue
        parts = xlrd.xldate_as_tuple(raw_date, book.datemode)
        observations.append((date(*parts[:3]), float(raw_price)))

    if not observations:
        raise ValueError("EIA 휘발유 이력이 비어 있습니다")
    return sorted(observations)


def parse_electricity_monthly(body: bytes) -> dict[str, tuple[float, str]]:
    """월간 전력요금 이력 → {YYYY-MM: (¢/kWh, 확정상태)} (뉴욕, 교통 부문).

    확정 상태를 함께 돌려주는 이유는 EIA 가 최근 약 17개월을 `Preliminary` 로 두고
    나중에 `Final` 로 바꾸기 때문입니다. 같은 달을 다시 만들었을 때 숫자가 달라지는
    유일한 원인이라, 결과에 남겨두면 그 차이를 설명할 수 있습니다.
    """
    workbook = openpyxl.load_workbook(BytesIO(body), read_only=True, data_only=True)
    if ELECTRICITY_SHEET not in workbook.sheetnames:
        raise ValueError(f"EIA 전력 파일에 {ELECTRICITY_SHEET} 시트가 없습니다")
    sheet = workbook[ELECTRICITY_SHEET]

    rows = sheet.iter_rows(values_only=True)
    # 0행 부문(병합 셀이라 앞으로 채움), 1행 항목, 2행 키 이름.
    sector_row, field_row, key_row = (next(rows) for _ in range(3))
    sectors, current = [], ""
    for value in sector_row:
        current = str(value).strip() if value not in (None, "") else current
        sectors.append(current)
    columns = [
        f"{sector}_{str(field).strip()}" if field not in (None, "") else str(key).strip()
        for sector, field, key in zip(sectors, field_row, key_row)
    ]

    price_column = f"{ELECTRICITY_SECTOR}_Price"
    for required in ("Year", "Month", "State", STATUS_COLUMN, price_column):
        if required not in columns:
            raise ValueError(f"EIA 전력 시트에 컬럼이 없습니다: {required}")
    index = {name: position for position, name in enumerate(columns)}

    prices: dict[str, tuple[float, str]] = {}
    for row in rows:
        if row[index["State"]] != STATE:
            continue
        price = row[index[price_column]]
        if not isinstance(price, (int, float)):
            continue
        year_month = f"{int(row[index['Year']]):04d}-{int(row[index['Month']]):02d}"
        status = str(row[index[STATUS_COLUMN]] or "").strip()
        prices[year_month] = (float(price), status)

    workbook.close()
    if not prices:
        raise ValueError(f"EIA 전력 이력에 {STATE} 데이터가 없습니다")
    return prices


def _gas_price_for(days: list[date], weekly: list[tuple[date, float]]) -> dict[date, float]:
    """각 날짜에 **그 날 이하 가장 최근 주간 관측치**를 복제합니다.

    선형 보간하지 않는 이유는, EIA 주간값이 "그 주의 관측 평균"이라 다음 관측까지
    유효한 값으로 보는 편이 원 데이터에 가깝기 때문입니다.
    """
    prices: dict[date, float] = {}
    for day in days:
        earlier = [price for observed, price in weekly if observed <= day]
        if not earlier:
            raise ValueError(
                f"{day} 이전의 휘발유 관측치가 없습니다 (원본 시작일 {weekly[0][0]} 이후여야 함)"
            )
        prices[day] = earlier[-1]
    return prices


def build_daily_prices(
    year_month: str,
    gas_body: bytes,
    electricity_body: bytes,
    bronze_collected_date: date,
    markup: float = PUBLIC_CHARGING_MARKUP,
) -> list[dict]:
    """대상 월의 일별 가격 + 계보 행.

    `bronze_collected_date` 를 받는 이유는 같은 달을 다시 만들면 숫자가 달라질 수 있어서
    입니다. 어느 수집분으로 만들었는지 남기지 않으면 그 차이를 설명할 수 없습니다.
    """
    datetime.strptime(year_month, "%Y-%m")
    days = month_days(year_month)

    gas_prices = _gas_price_for(days, parse_gas_weekly(gas_body))

    electricity = parse_electricity_monthly(electricity_body)
    if year_month not in electricity:
        available = f"{min(electricity)} ~ {max(electricity)}"
        raise ValueError(
            f"EIA 전력 이력에 {year_month} 이 없습니다 (보유 {available}). "
            "전력 통계는 약 3개월 늦게 공개됩니다."
        )
    cents, status = electricity[year_month]
    ev_price = cents / CENTS_PER_DOLLAR * markup

    rows = [
        {
            "date": day,
            "gas_price": gas_prices[day],
            "ev_price": ev_price,
            "price_source": PRICE_SOURCE,
            "bronze_collected_date": bronze_collected_date,
            "ev_price_status": status,
        }
        for day in days
    ]
    validate(rows, year_month)
    logger.info(
        "EIA 일별 연료비 생성: %s %d일 gas=%.3f~%.3f ev=%.4f (배수 %.2f) "
        "수집분=%s 전력상태=%s",
        year_month, len(rows),
        min(row["gas_price"] for row in rows), max(row["gas_price"] for row in rows),
        ev_price, markup, bronze_collected_date, status or "(표기없음)",
    )
    if status != FINAL:
        # 잠정값이면 나중에 다시 만들 때 숫자가 바뀝니다. 조용히 넘기지 않습니다.
        logger.warning(
            "%s 전력값이 확정(%s) 이 아닙니다 (%s). 나중에 다시 만들면 값이 바뀝니다.",
            year_month, FINAL, status or "표기없음",
        )
    return rows


def validate(rows: list[dict], year_month: str) -> None:
    """날짜 완결성과 가격 범위. 둘 다 하류에서 조용히 틀리는 경로라 여기서 막습니다.

    날짜가 하루라도 비면 Gold 의 일자 조인에서 그 날 운행이 통째로 매칭 실패합니다.
    Gold 는 매칭 실패를 예외로 잡지만, 그 메시지는 차량 문제와 구분되지 않아 원인을
    찾는 데 시간이 걸립니다. 여기서 막는 편이 훨씬 빠릅니다.
    """
    expected = month_days(year_month)
    if [row["date"] for row in rows] != expected:
        raise ValueError(f"{year_month} 일별 날짜가 그 달 전체와 다릅니다")

    for row in rows:
        if not GAS_USD_RANGE[0] < row["gas_price"] < GAS_USD_RANGE[1]:
            raise ValueError(f"휘발유 가격이 허용 범위 밖입니다: {row['gas_price']}")
        if not EV_USD_RANGE[0] < row["ev_price"] < EV_USD_RANGE[1]:
            raise ValueError(f"전기 단가가 허용 범위 밖입니다: {row['ev_price']}")
