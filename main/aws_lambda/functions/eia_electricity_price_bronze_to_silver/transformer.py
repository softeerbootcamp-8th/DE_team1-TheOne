"""EIA 월간 전력요금 이력에서 대상 월을 뽑아 일별 충전 단가로 펼칩니다.

원본은 **월 단위 ¢/kWh** 인데 출력은 **일별** 입니다(`CLEAN_EV_CHARGING_PRICE_SCHEMA`).
Gold 가 운행 날짜로 조인하므로 그 달 전 일수가 빠짐없이 있어야 하고, 하루라도 비면
그 날 운행이 통째로 매칭에 실패합니다 — 에러가 아니라 조용히 줄어든 집계로 나타납니다.
그래서 같은 월값을 그 달 모든 날에 채웁니다.

공공 충전 배수를 곱하는 이유
--------------------------
EIA 가 주는 값은 **교통 부문 전력 소매가** 입니다. 기사가 실제로 내는 공공 급속충전
요금은 그보다 비싸서, 배수를 곱해 실사용가에 맞춥니다. 이 값은 실측이 아니라 가정이라
`markup` 으로 열어 둡니다.
"""

import calendar
import logging
from datetime import date, datetime
from io import BytesIO

import openpyxl

logger = logging.getLogger(__name__)

ELECTRICITY_SHEET = "Monthly-States"
STATE = "NY"
ELECTRICITY_SECTOR = "TRANSPORTATION"
STATUS_COLUMN = "Data Status"
PUBLIC_CHARGING_MARKUP = 2.0
CENTS_PER_DOLLAR = 100.0
EV_USD_RANGE = (0.05, 3.0)
FINAL = "Final"


def month_days(year_month: str) -> list[date]:
    year, month = (int(part) for part in year_month.split("-"))
    last = calendar.monthrange(year, month)[1]
    return [date(year, month, day) for day in range(1, last + 1)]


def parse_electricity_monthly(body: bytes) -> dict[str, tuple[float, str]]:
    """월간 전력요금 이력 → {YYYY-MM: (¢/kWh, 확정상태)} (뉴욕, 교통 부문).

    확정 상태를 함께 돌려주는 이유는 EIA 가 최근 약 17개월을 `Preliminary` 로 두고
    나중에 `Final` 로 바꾸기 때문입니다. 같은 달을 다시 만들었을 때 숫자가 달라지는
    유일한 원인이라, 로그로 남겨두면 그 차이를 설명할 수 있습니다.
    """
    workbook = openpyxl.load_workbook(BytesIO(body), read_only=True, data_only=True)
    if ELECTRICITY_SHEET not in workbook.sheetnames:
        raise ValueError(f"EIA 전력 파일에 {ELECTRICITY_SHEET} 시트가 없습니다")
    sheet = workbook[ELECTRICITY_SHEET]

    rows = sheet.iter_rows(values_only=True)
    # 0행 부문(병합 셀이라 앞으로 채움), 1행 항목, 2행 키 이름.
    sector_row, field_row, key_row = (next(rows) for _ in range(3))
    sectors, current = [], ""
    for value in sector_row:
        current = str(value).strip() if value not in (None, "") else current
        sectors.append(current)
    columns = [
        f"{sector}_{str(field).strip()}" if field not in (None, "") else str(key).strip()
        for sector, field, key in zip(sectors, field_row, key_row)
    ]

    price_column = f"{ELECTRICITY_SECTOR}_Price"
    for required in ("Year", "Month", "State", STATUS_COLUMN, price_column):
        if required not in columns:
            raise ValueError(f"EIA 전력 시트에 컬럼이 없습니다: {required}")
    index = {name: position for position, name in enumerate(columns)}

    prices: dict[str, tuple[float, str]] = {}
    for row in rows:
        if row[index["State"]] != STATE:
            continue
        price = row[index[price_column]]
        if not isinstance(price, (int, float)):
            continue
        year_month = f"{int(row[index['Year']]):04d}-{int(row[index['Month']]):02d}"
        status = str(row[index[STATUS_COLUMN]] or "").strip()
        prices[year_month] = (float(price), status)

    workbook.close()
    if not prices:
        raise ValueError(f"EIA 전력 이력에 {STATE} 데이터가 없습니다")
    return prices


def validate(rows: list[dict], year_month: str) -> None:
    """그 달 전 일수가 빠짐없이 있고 단가가 허용 범위인지 봅니다."""
    expected = month_days(year_month)
    if [row["date"] for row in rows] != expected:
        raise ValueError(
            f"{year_month} 일자가 빠짐없이 있어야 합니다: "
            f"{len(rows)}행 (기대 {len(expected)}행)"
        )
    for row in rows:
        if not EV_USD_RANGE[0] < row["ev_price"] < EV_USD_RANGE[1]:
            raise ValueError(f"충전 단가가 허용 범위 밖입니다: {row['ev_price']}")


def build_daily_prices(
    year_month: str,
    electricity_body: bytes,
    bronze_collected_date: date,
    markup: float = PUBLIC_CHARGING_MARKUP,
) -> list[dict]:
    """대상 월의 일별 충전 단가."""
    datetime.strptime(year_month, "%Y-%m")

    electricity = parse_electricity_monthly(electricity_body)
    if year_month not in electricity:
        available = f"{min(electricity)} ~ {max(electricity)}"
        raise ValueError(
            f"EIA 전력 이력에 {year_month} 이 없습니다 (보유 {available}). "
            "전력 통계는 약 3개월 늦게 공개됩니다."
        )
    cents, status = electricity[year_month]
    ev_price = cents / CENTS_PER_DOLLAR * markup

    rows = [{"date": day, "ev_price": ev_price} for day in month_days(year_month)]
    validate(rows, year_month)

    logger.info(
        "EIA 일별 충전 단가 생성: %s %d일 ev=%.4f (배수 %.2f) 수집분=%s 전력상태=%s",
        year_month, len(rows), ev_price, markup, bronze_collected_date,
        status or "(표기없음)",
    )
    if status != FINAL:
        # 잠정값이면 나중에 다시 만들 때 숫자가 바뀝니다. 조용히 넘기지 않습니다.
        logger.warning(
            "%s 전력값이 확정(%s) 이 아닙니다 (%s). 나중에 다시 만들면 값이 바뀝니다.",
            year_month, FINAL, status or "표기없음",
        )
    return rows
